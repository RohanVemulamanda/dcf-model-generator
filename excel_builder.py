"""
excel_builder.py
--------------------------------
Phase 2b of the "AI DCF model generator" project.

Takes the outputs of dcf_engine.py for ANY company and writes an actual,
formula-driven Excel workbook -- not a static report. Every projection and
valuation cell is a real Excel formula that recalculates if you change an
input, exactly like the JNJ model. Tab structure mirrors that model:

    Cover | Historicals | WACC | Forecast | DCF | Sensitivity | Audit

Formatting follows standard IB convention:
    blue  = hardcoded input (something a human typed in / pulled from a filing)
    black = a formula that only uses cells on its own sheet
    green = a formula that pulls a value from a DIFFERENT sheet

Run `scripts/recalc.py` (from the xlsx skill) on the output file after this
script -- openpyxl writes formula TEXT but not their computed values, so
the file needs one recalculation pass before the numbers will show up in
Excel/Google Sheets previews.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF", size=10)          # hardcoded input
BLACK = Font(name=FONT_NAME, color="000000", size=10)          # same-sheet formula
GREEN = Font(name=FONT_NAME, color="008000", size=10)          # cross-sheet formula
BOLD_BLACK = Font(name=FONT_NAME, color="000000", size=10, bold=True)
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, color="000000", size=14, bold=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = "$#,##0;($#,##0);\"-\""
PCT_FMT = "0.0%;(0.0%);\"-\""
PRICE_FMT = "$#,##0.00"
MULTIPLE_FMT = "0.00x"


def _header_row(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD_BLACK if bold else BLACK
    return cell


def _write(ws, row, col, value, font, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if number_format:
        cell.number_format = number_format
    cell.border = BORDER
    return cell


def build_workbook(
    ticker: str,
    company_name: str,
    historicals: pd.DataFrame,     # rows = line items, columns = fiscal years (from sec_financials_fetcher)
    forecast: pd.DataFrame,        # output of dcf_engine.generate_forecast (values, not formulas)
    tax_rate: float,
    wacc_inputs,                   # dcf_engine.WaccInputs
    terminal_growth: float,
    total_debt: float,
    cash: float,
    shares_outstanding: float,
    market_price: float | None,
    output_path: str,
):
    wb = Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet("Cover")
    ws_hist = wb.create_sheet("Historicals")
    ws_wacc = wb.create_sheet("WACC")
    ws_fcst = wb.create_sheet("Forecast")
    ws_dcf = wb.create_sheet("DCF")
    ws_sens = wb.create_sheet("Sensitivity")
    ws_audit = wb.create_sheet("Audit")

    hist_years = list(historicals.columns)
    n_hist = len(hist_years)
    n_fcst = len(forecast)

    # =======================================================================
    # HISTORICALS
    # =======================================================================
    ws_hist["A1"] = f"{company_name} ({ticker}) — Historical Financials"
    ws_hist["A1"].font = TITLE_FONT
    _header_row(ws_hist, 3, ["Line Item ($mm)"] + [str(y) for y in hist_years])

    hist_row_of = {}  # line item -> Excel row number, so Forecast/WACC can reference it
    row = 4
    for line_item in historicals.index:
        _label(ws_hist, row, 1, line_item)
        for j, yr in enumerate(hist_years):
            val = historicals.loc[line_item, yr]
            val = None if pd.isna(val) else float(val)
            _write(ws_hist, row, 2 + j, val, BLUE, CURRENCY_FMT)
        hist_row_of[line_item] = row
        row += 1

    # Derived row: Gross Profit (formula, black -- same sheet)
    if "Revenue" in hist_row_of and "Cost of Goods Sold" in hist_row_of:
        gp_row = row
        _label(ws_hist, gp_row, 1, "Gross Profit", bold=True)
        rev_r, cogs_r = hist_row_of["Revenue"], hist_row_of["Cost of Goods Sold"]
        for j in range(n_hist):
            col_letter = get_column_letter(2 + j)
            _write(ws_hist, gp_row, 2 + j,
                   f"={col_letter}{rev_r}-{col_letter}{cogs_r}", BLACK, CURRENCY_FMT)
        hist_row_of["Gross Profit"] = gp_row
        row += 1

    ws_hist.column_dimensions["A"].width = 30
    for j in range(n_hist):
        ws_hist.column_dimensions[get_column_letter(2 + j)].width = 14
    src_row = row + 1
    ws_hist.cell(row=src_row, column=1,
                 value="Source: SEC EDGAR, XBRL company facts (10-K filings), pulled via sec_financials_fetcher.py").font = Font(
        name=FONT_NAME, size=8, italic=True)

    last_hist_col_letter = get_column_letter(2 + n_hist - 1)

    # =======================================================================
    # WACC
    # =======================================================================
    ws_wacc["A1"] = f"{company_name} ({ticker}) — WACC (CAPM)"
    ws_wacc["A1"].font = TITLE_FONT

    labels_inputs = [
        ("Risk-Free Rate (10-Yr UST)", wacc_inputs.risk_free_rate, PCT_FMT, "Source: FRED, series DGS10"),
        ("Equity Risk Premium", wacc_inputs.equity_risk_premium, PCT_FMT, "Source: A. Damodaran (NYU Stern), implied ERP"),
        ("Relevered Beta", wacc_inputs.beta, "0.000", "Source: A. Damodaran industry beta, relevered to company D/E"),
        ("Pre-Tax Cost of Debt", wacc_inputs.pretax_cost_of_debt, PCT_FMT, "Source: A. Damodaran, industry avg cost of debt"),
        ("Tax Rate", tax_rate, PCT_FMT, "Marginal tax rate assumption"),
        ("Market Value of Equity ($mm)", wacc_inputs.market_value_equity, CURRENCY_FMT, "Shares outstanding x market price"),
        ("Total Debt ($mm)", wacc_inputs.total_debt, CURRENCY_FMT, f"Source: {company_name} 10-K, most recent balance sheet"),
    ]
    row = 3
    wacc_row_of = {}
    for label, value, fmt, note in labels_inputs:
        _label(ws_wacc, row, 1, label)
        cell = _write(ws_wacc, row, 2, value, BLUE, fmt)
        cell.fill = ASSUMPTION_FILL
        ws_wacc.cell(row=row, column=4, value=note).font = Font(name=FONT_NAME, size=8, italic=True)
        wacc_row_of[label] = row
        row += 1

    row += 1
    _label(ws_wacc, row, 1, "Cost of Equity (CAPM)", bold=True)
    coe_row = row
    rf_r, beta_r, erp_r = wacc_row_of["Risk-Free Rate (10-Yr UST)"], wacc_row_of["Relevered Beta"], wacc_row_of["Equity Risk Premium"]
    _write(ws_wacc, row, 2, f"=B{rf_r}+B{beta_r}*B{erp_r}", BLACK, PCT_FMT)
    row += 1

    _label(ws_wacc, row, 1, "After-Tax Cost of Debt", bold=True)
    atkd_row = row
    kd_r, tax_r = wacc_row_of["Pre-Tax Cost of Debt"], wacc_row_of["Tax Rate"]
    _write(ws_wacc, row, 2, f"=B{kd_r}*(1-B{tax_r})", BLACK, PCT_FMT)
    row += 1

    _label(ws_wacc, row, 1, "Weight of Equity", bold=True)
    we_row = row
    mve_r, debt_r = wacc_row_of["Market Value of Equity ($mm)"], wacc_row_of["Total Debt ($mm)"]
    _write(ws_wacc, row, 2, f"=B{mve_r}/(B{mve_r}+B{debt_r})", BLACK, PCT_FMT)
    row += 1

    _label(ws_wacc, row, 1, "Weight of Debt", bold=True)
    wd_row = row
    _write(ws_wacc, row, 2, f"=B{debt_r}/(B{mve_r}+B{debt_r})", BLACK, PCT_FMT)
    row += 2

    _label(ws_wacc, row, 1, "WACC", bold=True)
    wacc_final_row = row
    cell = _write(ws_wacc, row, 2, f"=B{we_row}*B{coe_row}+B{wd_row}*B{atkd_row}", BLACK, PCT_FMT)
    cell.font = Font(name=FONT_NAME, color="000000", size=11, bold=True)

    ws_wacc.column_dimensions["A"].width = 30
    ws_wacc.column_dimensions["B"].width = 14
    ws_wacc.column_dimensions["D"].width = 55

    WACC_CELL = f"WACC!$B${wacc_final_row}"
    TAX_CELL = f"WACC!$B${tax_r}"

    # =======================================================================
    # FORECAST
    # =======================================================================
    ws_fcst["A1"] = f"{company_name} ({ticker}) — 5-Year Forecast"
    ws_fcst["A1"].font = TITLE_FONT
    _header_row(ws_fcst, 3, ["Line Item"] + list(forecast["Year"]))

    fcst_row_of = {}
    row = 4

    def fcst_input_row(label, series, fmt):
        nonlocal row
        _label(ws_fcst, row, 1, label)
        for j in range(n_fcst):
            cell = _write(ws_fcst, row, 2 + j, float(series.iloc[j]), BLUE, fmt)
            cell.fill = ASSUMPTION_FILL
        fcst_row_of[label] = row
        row += 1

    fcst_input_row("Revenue Growth %", forecast["Revenue Growth %"], PCT_FMT)

    rev_row = row
    _label(ws_fcst, row, 1, "Revenue", bold=True)
    growth_r = fcst_row_of["Revenue Growth %"]
    last_actual_rev_cell = f"Historicals!{last_hist_col_letter}{hist_row_of['Revenue']}"
    for j in range(n_fcst):
        col_letter = get_column_letter(2 + j)
        if j == 0:
            formula = f"={last_actual_rev_cell}*(1+{col_letter}{growth_r})"
            font = GREEN
        else:
            prev_col = get_column_letter(2 + j - 1)
            formula = f"={prev_col}{rev_row}*(1+{col_letter}{growth_r})"
            font = BLACK
        _write(ws_fcst, row, 2 + j, formula, font, CURRENCY_FMT)
    fcst_row_of["Revenue"] = rev_row
    row += 1

    fcst_input_row("Gross Margin %", forecast["Gross Margin %"], PCT_FMT)
    fcst_input_row("SG&A % of Revenue", forecast["SG&A % of Revenue"], PCT_FMT)
    fcst_input_row("R&D % of Revenue", forecast["R&D % of Revenue"], PCT_FMT)
    fcst_input_row("D&A % of Revenue", forecast["D&A % of Revenue"], PCT_FMT)
    fcst_input_row("Capex % of Revenue", forecast["Capex % of Revenue"], PCT_FMT)
    fcst_input_row("NWC % of Revenue", forecast["NWC % of Revenue"], PCT_FMT)

    def derived_row(label, build_formula, fmt=CURRENCY_FMT, bold=False):
        nonlocal row
        _label(ws_fcst, row, 1, label, bold=bold)
        for j in range(n_fcst):
            col_letter = get_column_letter(2 + j)
            _write(ws_fcst, row, 2 + j, build_formula(col_letter, j), BLACK, fmt)
        fcst_row_of[label] = row
        row += 1

    gm_r = fcst_row_of["Gross Margin %"]
    derived_row("Gross Profit", lambda c, j: f"={c}{rev_row}*{c}{gm_r}")
    gp_r = fcst_row_of["Gross Profit"]

    sga_pct_r = fcst_row_of["SG&A % of Revenue"]
    derived_row("SG&A", lambda c, j: f"={c}{rev_row}*{c}{sga_pct_r}")
    sga_r = fcst_row_of["SG&A"]

    rd_pct_r = fcst_row_of["R&D % of Revenue"]
    derived_row("R&D", lambda c, j: f"={c}{rev_row}*{c}{rd_pct_r}")
    rd_r = fcst_row_of["R&D"]

    derived_row("EBIT", lambda c, j: f"={c}{gp_r}-{c}{sga_r}-{c}{rd_r}", bold=True)
    ebit_r = fcst_row_of["EBIT"]

    derived_row("NOPAT", lambda c, j: f"={c}{ebit_r}*(1-{TAX_CELL})")
    nopat_r = fcst_row_of["NOPAT"]
    for j in range(n_fcst):
        ws_fcst.cell(row=nopat_r, column=2 + j).font = GREEN  # references WACC tab's tax rate

    da_pct_r = fcst_row_of["D&A % of Revenue"]
    derived_row("D&A", lambda c, j: f"={c}{rev_row}*{c}{da_pct_r}")
    da_r = fcst_row_of["D&A"]

    capex_pct_r = fcst_row_of["Capex % of Revenue"]
    derived_row("Capex", lambda c, j: f"={c}{rev_row}*{c}{capex_pct_r}")
    capex_r = fcst_row_of["Capex"]

    nwc_pct_r = fcst_row_of["NWC % of Revenue"]
    derived_row("NWC Balance", lambda c, j: f"={c}{rev_row}*{c}{nwc_pct_r}")
    nwc_bal_r = fcst_row_of["NWC Balance"]

    # Change in NWC needs the prior period's balance; year 1 needs the LAST
    # ACTUAL historical NWC, which we compute inline from Historicals if the
    # working-capital rows exist, else 0.
    change_row = row
    _label(ws_fcst, row, 1, "Change in NWC")
    have_wc_rows = all(k in hist_row_of for k in ["Accounts Receivable", "Inventory", "Accounts Payable"])
    if have_wc_rows:
        ar_r, inv_r, ap_r = hist_row_of["Accounts Receivable"], hist_row_of["Inventory"], hist_row_of["Accounts Payable"]
        prior_nwc_actual = (f"(Historicals!{last_hist_col_letter}{ar_r}"
                             f"+Historicals!{last_hist_col_letter}{inv_r}"
                             f"-Historicals!{last_hist_col_letter}{ap_r})")
    else:
        prior_nwc_actual = "0"
    for j in range(n_fcst):
        col_letter = get_column_letter(2 + j)
        if j == 0:
            formula = f"={col_letter}{nwc_bal_r}-{prior_nwc_actual}"
            font = GREEN if have_wc_rows else BLACK
        else:
            prev_col = get_column_letter(2 + j - 1)
            formula = f"={col_letter}{nwc_bal_r}-{prev_col}{nwc_bal_r}"
            font = BLACK
        _write(ws_fcst, row, 2 + j, formula, font, CURRENCY_FMT)
    fcst_row_of["Change in NWC"] = change_row
    row += 1

    derived_row("FCFF", lambda c, j: f"={c}{nopat_r}+{c}{da_r}-{c}{capex_r}-{c}{change_row}", bold=True)
    fcff_r = fcst_row_of["FCFF"]

    ws_fcst.column_dimensions["A"].width = 26
    for j in range(n_fcst):
        ws_fcst.column_dimensions[get_column_letter(2 + j)].width = 13

    # =======================================================================
    # DCF
    # =======================================================================
    ws_dcf["A1"] = f"{company_name} ({ticker}) — DCF Valuation"
    ws_dcf["A1"].font = TITLE_FONT
    _header_row(ws_dcf, 3, ["Line Item"] + list(forecast["Year"]))

    row = 4
    _label(ws_dcf, row, 1, "FCFF")
    fcff_link_row = row
    for j in range(n_fcst):
        col_letter = get_column_letter(2 + j)
        _write(ws_dcf, row, 2 + j, f"=Forecast!{col_letter}{fcff_r}", GREEN, CURRENCY_FMT)
    row += 1

    _label(ws_dcf, row, 1, "Discount Factor")
    df_row = row
    for j in range(n_fcst):
        col_letter = get_column_letter(2 + j)
        _write(ws_dcf, row, 2 + j, f"=1/(1+{WACC_CELL})^{j + 1}", GREEN, "0.000")
    row += 1

    _label(ws_dcf, row, 1, "PV of FCFF", bold=True)
    pv_row = row
    for j in range(n_fcst):
        col_letter = get_column_letter(2 + j)
        _write(ws_dcf, row, 2 + j, f"={col_letter}{fcff_link_row}*{col_letter}{df_row}", BLACK, CURRENCY_FMT)
    row += 2

    last_col_letter = get_column_letter(2 + n_fcst - 1)

    _label(ws_dcf, row, 1, "Terminal Growth Rate")
    g_row = row
    cell = _write(ws_dcf, row, 2, terminal_growth, BLUE, PCT_FMT)
    cell.fill = ASSUMPTION_FILL
    row += 1

    _label(ws_dcf, row, 1, "Terminal Value")
    tv_row = row
    _write(ws_dcf, row, 2,
           f"=({last_col_letter}{fcff_link_row}*(1+B{g_row}))/({WACC_CELL}-B{g_row})", GREEN, CURRENCY_FMT)
    row += 1

    _label(ws_dcf, row, 1, "PV of Terminal Value")
    pv_tv_row = row
    _write(ws_dcf, row, 2, f"=B{tv_row}*{last_col_letter}{df_row}", BLACK, CURRENCY_FMT)
    row += 2

    _label(ws_dcf, row, 1, "Sum of PV of FCFF")
    sum_pv_row = row
    _write(ws_dcf, row, 2, f"=SUM(B{pv_row}:{last_col_letter}{pv_row})", BLACK, CURRENCY_FMT)
    row += 1

    _label(ws_dcf, row, 1, "Enterprise Value", bold=True)
    ev_row = row
    cell = _write(ws_dcf, row, 2, f"=B{sum_pv_row}+B{pv_tv_row}", BLACK, CURRENCY_FMT)
    cell.font = Font(name=FONT_NAME, size=11, bold=True)
    row += 2

    _label(ws_dcf, row, 1, "Less: Total Debt")
    debt_row = row
    cell = _write(ws_dcf, row, 2, total_debt, BLUE, CURRENCY_FMT)
    cell.fill = ASSUMPTION_FILL
    row += 1

    _label(ws_dcf, row, 1, "Plus: Cash & Equivalents")
    cash_row = row
    cell = _write(ws_dcf, row, 2, cash, BLUE, CURRENCY_FMT)
    cell.fill = ASSUMPTION_FILL
    row += 1

    _label(ws_dcf, row, 1, "Equity Value", bold=True)
    eq_row = row
    cell = _write(ws_dcf, row, 2, f"=B{ev_row}-B{debt_row}+B{cash_row}", BLACK, CURRENCY_FMT)
    cell.font = Font(name=FONT_NAME, size=11, bold=True)
    row += 1

    _label(ws_dcf, row, 1, "Diluted Shares Outstanding (mm)")
    shares_row = row
    cell = _write(ws_dcf, row, 2, shares_outstanding, BLUE, "#,##0.0")
    cell.fill = ASSUMPTION_FILL
    row += 1

    _label(ws_dcf, row, 1, "Implied Value per Share", bold=True)
    price_row = row
    cell = _write(ws_dcf, row, 2, f"=B{eq_row}/B{shares_row}", BLACK, PRICE_FMT)
    cell.font = Font(name=FONT_NAME, size=12, bold=True)
    row += 1

    if market_price is not None:
        _label(ws_dcf, row, 1, "Current Market Price")
        mkt_row = row
        _write(ws_dcf, row, 2, market_price, BLUE, PRICE_FMT)
        row += 1
        _label(ws_dcf, row, 1, "Implied Upside / (Downside)", bold=True)
        _write(ws_dcf, row, 2, f"=B{price_row}/B{mkt_row}-1", BLACK, PCT_FMT)
        row += 1

    ws_dcf.column_dimensions["A"].width = 28
    for j in range(n_fcst):
        ws_dcf.column_dimensions[get_column_letter(2 + j)].width = 13

    # =======================================================================
    # SENSITIVITY  (WACC down the side, terminal growth across the top)
    # =======================================================================
    ws_sens["A1"] = f"{company_name} ({ticker}) — Sensitivity: Implied Share Price"
    ws_sens["A1"].font = TITLE_FONT
    ws_sens["A2"] = "WACC (rows) vs. Terminal Growth Rate (columns)"
    ws_sens["A2"].font = Font(name=FONT_NAME, size=9, italic=True)

    grid_size = 5
    wacc_step, g_step = 0.005, 0.0025
    half = grid_size // 2

    header_row_num = 4
    ws_sens.cell(row=header_row_num, column=1, value="WACC \\ g").font = BOLD_BLACK
    for k in range(grid_size):
        col = 2 + k
        offset = k - half
        cell = ws_sens.cell(row=header_row_num, column=col, value=f"={WACC_CELL.replace('WACC!', 'DCF!B' + str(g_row) + '+0*')}")
        # Column header = terminal growth values, centered on DCF!g
        cell.value = f"=DCF!$B${g_row}+{offset}*{g_step}"
        cell.font = GREEN
        cell.number_format = PCT_FMT
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT_NAME, color="FFFFFF", bold=True)

    fcff_cells = [f"Forecast!{get_column_letter(2 + j)}{fcff_r}" for j in range(n_fcst)]

    for r in range(grid_size):
        row_num = header_row_num + 1 + r
        offset = r - half
        wcell = ws_sens.cell(row=row_num, column=1, value=f"={WACC_CELL}+{offset}*{wacc_step}")
        wcell.font = GREEN
        wcell.number_format = PCT_FMT
        wcell.fill = HEADER_FILL
        wcell.font = Font(name=FONT_NAME, color="FFFFFF", bold=True)

        for c in range(grid_size):
            col_num = 2 + c
            col_letter = get_column_letter(col_num)
            w_ref = f"$A{row_num}"
            g_ref = f"{col_letter}${header_row_num}"

            pv_terms = "+".join(
                f"{fcff_cells[j]}/(1+{w_ref})^{j + 1}" for j in range(n_fcst)
            )
            tv_term = f"({fcff_cells[-1]}*(1+{g_ref}))/({w_ref}-{g_ref})"
            pv_tv_term = f"{tv_term}/(1+{w_ref})^{n_fcst}"
            ev_term = f"({pv_terms}+{pv_tv_term})"
            price_formula = f"=IF({w_ref}<={g_ref},\"n/a\",({ev_term}-{total_debt}+{cash})/{shares_outstanding})"

            cell = ws_sens.cell(row=row_num, column=col_num, value=price_formula)
            cell.font = BLACK
            cell.number_format = PRICE_FMT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    ws_sens.column_dimensions["A"].width = 12
    for c in range(grid_size):
        ws_sens.column_dimensions[get_column_letter(2 + c)].width = 12

    # =======================================================================
    # AUDIT
    # =======================================================================
    ws_audit["A1"] = f"{company_name} ({ticker}) — Audit Checks"
    ws_audit["A1"].font = TITLE_FONT
    _header_row(ws_audit, 3, ["Check", "Result", "Detail"])

    audit_checks = [
        ("WACC within 3%-15%", f'=IF(AND({WACC_CELL}>=0.03,{WACC_CELL}<=0.15),"PASS","FAIL")',
         f'=TEXT({WACC_CELL},"0.0%")'),
        ("WACC exceeds terminal growth rate", f"=IF({WACC_CELL}>DCF!B{g_row},\"PASS\",\"FAIL\")",
         f'=TEXT({WACC_CELL}-DCF!B{g_row},"0.00%")&" spread"'),
        ("Capital structure weights sum to 100%",
         f"=IF(ABS((WACC!B{we_row}+WACC!B{wd_row})-1)<0.001,\"PASS\",\"FAIL\")", ""),
        ("Enterprise value is positive", f"=IF(DCF!B{ev_row}>0,\"PASS\",\"FAIL\")",
         f'=TEXT(DCF!B{ev_row},"$#,##0")&"mm"'),
        ("Equity value is positive", f"=IF(DCF!B{eq_row}>0,\"PASS\",\"FAIL\")",
         f'=TEXT(DCF!B{eq_row},"$#,##0")&"mm"'),
        ("Implied share price is positive", f"=IF(DCF!B{price_row}>0,\"PASS\",\"FAIL\")",
         f'=TEXT(DCF!B{price_row},"$#,##0.00")'),
        ("Terminal value is under 90% of enterprise value",
         f"=IF((DCF!B{pv_tv_row}/DCF!B{ev_row})<0.9,\"PASS\",\"FAIL\")",
         f'=TEXT(DCF!B{pv_tv_row}/DCF!B{ev_row},"0.0%")&" of EV"'),
        ("No blank cells in the Forecast revenue row",
         f'=IF(COUNTBLANK(Forecast!B{rev_row}:{last_col_letter}{rev_row})=0,"PASS","FAIL")', ""),
    ]

    row = 4
    for name, result_formula, detail_formula in audit_checks:
        _label(ws_audit, row, 1, name)
        rcell = ws_audit.cell(row=row, column=2, value=result_formula)
        rcell.font = GREEN
        rcell.alignment = Alignment(horizontal="center")
        rcell.border = BORDER
        if detail_formula:
            dcell = ws_audit.cell(row=row, column=3, value=detail_formula)
            dcell.font = GREEN
            dcell.border = BORDER
        row += 1

    ws_audit.column_dimensions["A"].width = 42
    ws_audit.column_dimensions["B"].width = 12
    ws_audit.column_dimensions["C"].width = 20

    # =======================================================================
    # COVER (written last so it can link to final row numbers everywhere)
    # =======================================================================
    ws_cover["A1"] = f"{company_name} ({ticker})"
    ws_cover["A1"].font = Font(name=FONT_NAME, size=18, bold=True)
    ws_cover["A2"] = "DCF Valuation Model (auto-generated)"
    ws_cover["A2"].font = Font(name=FONT_NAME, size=12, italic=True)

    summary = [
        ("WACC", f"=WACC!B{wacc_final_row}", PCT_FMT),
        ("Terminal Growth Rate", f"=DCF!B{g_row}", PCT_FMT),
        ("Enterprise Value ($mm)", f"=DCF!B{ev_row}", CURRENCY_FMT),
        ("Equity Value ($mm)", f"=DCF!B{eq_row}", CURRENCY_FMT),
        ("Implied Value per Share", f"=DCF!B{price_row}", PRICE_FMT),
    ]
    row = 5
    for label, formula, fmt in summary:
        _label(ws_cover, row, 1, label, bold=True)
        cell = _write(ws_cover, row, 2, formula, GREEN, fmt)
        cell.font = Font(name=FONT_NAME, color="008000", size=12, bold=True)
        row += 1

    row += 1
    ws_cover.cell(row=row, column=1,
                   value="Tabs: Historicals -> WACC -> Forecast -> DCF -> Sensitivity -> Audit").font = Font(
        name=FONT_NAME, size=9, italic=True)
    row += 1
    ws_cover.cell(row=row, column=1,
                   value="Font color key: blue = hardcoded input, black = same-sheet formula, green = cross-sheet link").font = Font(
        name=FONT_NAME, size=9, italic=True)

    ws_cover.column_dimensions["A"].width = 26
    ws_cover.column_dimensions["B"].width = 18

    wb.save(output_path)
    return output_path
